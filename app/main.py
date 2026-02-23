# Web app. Serves the UI, the analyze API, and the diff report download.

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from app.data_provider import fetch_prices
from app.engine import build_indicator_df, compute_sma, compute_ema

STATIC_DIR = Path(__file__).parent / "static"
VALIDATION_CSV = Path(__file__).resolve().parent.parent / "validation_mmm_full.csv"
EXCEL_WORKBOOK = Path(__file__).resolve().parent.parent / "example.xlsm"

app = FastAPI(title="SMA / EMA Analyzer", version="1.0.0")
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/", response_class=HTMLResponse)
async def index():
    return (STATIC_DIR / "index.html").read_text()


@app.get("/api/analyze")
async def analyze(
    ticker: str = Query("MMM", description="Stock ticker symbol"),
    start_date: str = Query("2016-01-01", description="Start date YYYY-MM-DD"),
    window: int = Query(10, ge=2, le=500, description="Moving average period"),
    mode: str = Query("SMA", description="SMA or EMA"),
):
    if mode.upper() not in ("SMA", "EMA"):
        return JSONResponse({"error": f"Unknown mode '{mode}'; expected 'SMA' or 'EMA'"}, status_code=400)

    ticker = ticker.strip()
    if not ticker:
        return JSONResponse({"error": "Ticker is required"}, status_code=400)

    try:
        price_df = fetch_prices(ticker, start_date)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    result = build_indicator_df(
        dates=price_df["Date"],
        prices=price_df["Close"],
        window=window,
        mode=mode,
    )

    # Convert to list of dicts for JSON. NaN -> None so it serializes cleanly.
    result["Date"] = result["Date"].dt.strftime("%Y-%m-%d")
    records = []
    for _, row in result.iterrows():
        close_val = row["Close"]
        ind_val = row["Indicator"]
        records.append({
            "Date": row["Date"],
            "Close": float(close_val) if pd.notna(close_val) else None,
            "Indicator": float(ind_val) if pd.notna(ind_val) else None,
            "Indicator_Label": row.get("Indicator_Label"),
        })

    return JSONResponse({
        "ticker": ticker.upper(),
        "mode": mode.upper(),
        "window": window,
        "start_date": start_date,
        "data": records,
    })


def _extract_excel_sma_data():
    # Pull dates, prices, and Excel's SMA values from the SMA_Graph sheet.
    import openpyxl

    wb = openpyxl.load_workbook(str(EXCEL_WORKBOOK), data_only=True, read_only=True)
    ws = wb["SMA_Graph"]

    window = 10
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=2, values_only=True):
        window = int(row[0])

    dates, prices, sma_vals = [], [], []
    for row in ws.iter_rows(min_row=7, max_row=9, values_only=False):
        label = str(row[0].value) if row[0].value else ""
        for c in row[1:]:
            val = c.value
            if val is None or str(val) == "#N/A":
                continue
            if "Trade Dates" in label and hasattr(val, "strftime"):
                dates.append(val.strftime("%Y-%m-%d"))
            elif label == "Price" and isinstance(val, (int, float)):
                prices.append(float(val))
            elif "Moving Avg" in label and isinstance(val, (int, float)):
                sma_vals.append(float(val))

    wb.close()
    return dates, prices, sma_vals, window


@app.get("/api/diff-report")
async def diff_report(
    ticker: str = Query("MMM"),
    start_date: str = Query("2016-01-01"),
    window: int = Query(10, ge=2, le=500),
    mode: str = Query("SMA"),
):
    # When mode is SMA and ticker is MMM, compare against Excel baseline.
    # For anything else, generate a standalone report with the computed values.
    mode = mode.upper()

    if mode == "SMA" and ticker.upper() == "MMM" and EXCEL_WORKBOOK.exists():
        dates_desc, prices_desc, excel_sma_desc, excel_window = _extract_excel_sma_data()

        dates_asc = list(reversed(dates_desc))
        prices_asc = list(reversed(prices_desc))

        close_series = pd.Series(prices_asc, dtype="float64")
        python_sma = compute_sma(close_series, excel_window)
        python_sma_desc = list(reversed(python_sma.tolist()))

        rows = []
        for i in range(len(excel_sma_desc)):
            excel_s = excel_sma_desc[i]
            python_s = python_sma_desc[i]
            if python_s is None or pd.isna(python_s):
                continue
            diff = abs(excel_s - python_s)
            rows.append({
                "Date": dates_desc[i],
                "Close": prices_desc[i],
                "Excel_SMA": excel_s,
                "Python_SMA": python_s,
                "Diff": diff,
            })

        report = pd.DataFrame(rows)
        max_diff = report["Diff"].abs().max() if len(report) else 0
        mean_diff = report["Diff"].abs().mean() if len(report) else 0

        buf = io.StringIO()
        buf.write(f"# Parity Report - MMM SMA({excel_window}) vs Excel\n")
        buf.write(f"# Max Abs Diff:  {max_diff:.2e}\n")
        buf.write(f"# Mean Abs Diff: {mean_diff:.2e}\n")
        buf.write(f"# Total Rows:    {len(report)}\n")
        buf.write(f"# Status:        {'PASS' if max_diff < 1e-6 else 'INVESTIGATE'}\n")
        buf.write("#\n")
        report.to_csv(buf, index=False)

    else:
        # For EMA or non-MMM tickers, pull live data and generate a report
        try:
            price_df = fetch_prices(ticker, start_date)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=400)

        close = price_df["Close"].astype(float)
        if mode == "EMA":
            indicator = compute_ema(close, window)
        else:
            indicator = compute_sma(close, window)

        report = pd.DataFrame({
            "Date": price_df["Date"].dt.strftime("%Y-%m-%d"),
            "Close": close.values,
            f"{mode}({window})": indicator.values,
        }).dropna(subset=[f"{mode}({window})"])

        buf = io.StringIO()
        buf.write(f"# {ticker.upper()} {mode}({window}) Report\n")
        buf.write(f"# Data points: {len(report)}\n")
        buf.write(f"# Date range:  {report['Date'].iloc[0]} to {report['Date'].iloc[-1]}\n")
        if mode == "EMA":
            buf.write(f"# Smoothing k: {2.0/(window+1):.6f}\n")
            buf.write(f"# Seed method: SMA of first {window} closes\n")
        buf.write("#\n")
        report.to_csv(buf, index=False)

    buf.seek(0)
    filename = f"{ticker.upper()}_{mode}{window}_report.csv"

    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
