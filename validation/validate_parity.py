# Compares our Python SMA to Excel's. Reads the workbook, pulls Excel's SMA values,
# computes ours from the same closes, and prints a diff. Run: python validation/validate_parity.py

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.engine import compute_sma, compute_ema


def extract_sma_graph_data(workbook_path: str):
    # Get dates, prices, and Excel's SMA from the SMA_Graph sheet.
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["SMA_Graph"]

    sma_window = None
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=2, values_only=True):
        sma_window = int(row[0])

    dates, prices, sma_vals = [], [], []

    for row in ws.iter_rows(min_row=7, max_row=9, values_only=False):
        label = str(row[0].value) if row[0].value else ""
        for c in row[1:]:
            val = c.value
            if val is None or str(val) == "#N/A":
                continue
            if "Trade Dates" in label and hasattr(val, "strftime"):
                dates.append(val.strftime("%Y-%m-%d"))
            elif label == "Price" and isinstance(val, (int, float)):
                prices.append(float(val))
            elif "Moving Avg" in label and isinstance(val, (int, float)):
                sma_vals.append(float(val))

    wb.close()
    return dates, prices, sma_vals, sma_window


def main():
    workbook = Path(__file__).resolve().parent.parent / "example.xlsm"
    if not workbook.exists():
        print(f"ERROR: {workbook} not found")
        sys.exit(1)

    dates, prices_desc, excel_sma_desc, window = extract_sma_graph_data(str(workbook))

    print(f"SMA Window: {window}")
    print(f"Dates extracted: {len(dates)}")
    print(f"Prices extracted: {len(prices_desc)}")
    print(f"Excel SMA values: {len(excel_sma_desc)}")

    # Excel is newest-first. Flip to chronological, run our engine, flip back to compare.
    dates_chr = list(reversed(dates))
    prices_chr = list(reversed(prices_desc))

    df = pd.DataFrame({"Date": pd.to_datetime(dates_chr), "Close": prices_chr})
    python_sma = compute_sma(df["Close"], window)
    python_sma_desc = list(reversed(python_sma.tolist()))

    valid_count = min(len(excel_sma_desc), len(python_sma_desc))

    diffs = []
    print(f"\n{'Date':<12} {'Excel_Close':>12} {'Excel_SMA':>12} {'Python_SMA':>12} {'Diff':>14}")
    print("-" * 65)

    for i in range(valid_count):
        excel_s = excel_sma_desc[i]
        python_s = python_sma_desc[i]
        if python_s is None or pd.isna(python_s):
            continue
        diff = abs(excel_s - python_s)
        diffs.append(diff)
        if i < 10 or i >= valid_count - 5:
            print(f"{dates[i]:<12} {prices_desc[i]:>12.4f} {excel_s:>12.6f} {python_s:>12.6f} {diff:>14.2e}")
        elif i == 10:
            print(f"{'...':^65}")

    max_diff = max(diffs) if diffs else 0
    mean_diff = sum(diffs) / len(diffs) if diffs else 0

    print(f"\n{'='*65}")
    print(f"PARITY SUMMARY")
    print(f"  Rows compared:  {len(diffs)}")
    print(f"  Max abs diff:   {max_diff:.2e}")
    print(f"  Mean abs diff:  {mean_diff:.2e}")

    if max_diff < 1e-12:
        print(f"  STATUS:         PASS (< 1e-12 threshold)")
    elif max_diff < 1e-6:
        print(f"  STATUS:         PASS (< 1e-6 threshold, floating-point noise)")
    else:
        print(f"  STATUS:         FAIL - investigate differences")

    out = Path(__file__).resolve().parent / "diff_report.csv"
    with open(out, "w") as f:
        f.write("Date,Excel_Close,Excel_SMA,Python_SMA,Diff\n")
        for i in range(valid_count):
            excel_s = excel_sma_desc[i]
            python_s = python_sma_desc[i]
            if python_s is None or pd.isna(python_s):
                continue
            diff = abs(excel_s - python_s)
            f.write(f"{dates[i]},{prices_desc[i]},{excel_s},{python_s},{diff}\n")

    print(f"\n  Report written to: {out}")

    # Also show EMA output (no Excel baseline for it, just documenting what we produce)
    print(f"\n{'='*65}")
    print("EMA VALIDATION (no Excel reference, just documenting what Python produces)")
    python_ema = compute_ema(df["Close"], window)
    ema_valid = python_ema.dropna()
    print(f"  EMA({window}) computed for {len(ema_valid)} data points")
    print(f"  First EMA value (seeded from SMA of first {window} prices): {ema_valid.iloc[0]:.6f}")
    print(f"  Smoothing factor k = 2/({window}+1) = {2.0/(window+1):.6f}")
    print(f"  Last 5 EMA values:")
    for i in range(-5, 0):
        d = df["Date"].iloc[i].strftime("%Y-%m-%d")
        p = df["Close"].iloc[i]
        e = python_ema.iloc[i]
        print(f"    {d}  Close={p:>10.4f}  EMA={e:>10.6f}")


if __name__ == "__main__":
    main()
